from paddleocr import PaddleOCR, draw_ocr
import pyautogui
import os
from PIL import Image
import time
import openpyxl
from openpyxl.reader.excel import load_workbook
import cv2
import winsound


excel_path = r'E:\python\xulie.xlsx'
base_path1 = r'E:\python\pic\function'
pyautogui.FAILSAFE = False
save_path1 = r'E:\python\pic\tech_current_name\screenshot1.png'
save_path2 = r'E:\python\pic\tech_current_name\screenshot2.png'

image_names = {'LEIDIAN': 'LEIDIAN.png', 'DIAOKEJI': 'DIAOKEJI.png', 'reserch': 'reserch.png', 'study': 'study.png',
               'body': 'body.png', 'accelerate': 'accelerate.png', 'accelerate_reserach': 'accelerate_reserach.png',
               'accelerate_reserach1': 'accelerate_reserach1.png', 'get_award': 'get_award.png',
               'quality': 'quality.png', 'temple': 'temple.png', 'forge_statue': 'forge_statue.png',
               'recycling_statues': 'recycling_statues.png', 'temple_button': 'temple_button.png',
               'recycling': 'recycling.png', 'double_click': 'double_click.png',}

region_names = {'GREEN_DIAOKEJI_NAME': (368, 268, 260, 48), 'GREEN_QUALITY_NAME': (445, 678, 100, 25),
                'BULE_DIAOKEJI_NAME': (720, 268, 260, 48), 'BULE_QUALITY_NAME': (725, 675, 136, 33),
                'RED_DIAOKEJI_NAME': (1010, 268, 260, 48), 'RED_QUALITY_NAME': (1020, 678, 110, 30),
                'PURPLE_DIAOKEJI_NAME': (1295, 268, 260, 48), 'PURPLE_QUALITY_NAME': (1316, 678, 110, 30),
                'FORGE_STATUE_NAME': (700, 339, 455, 85), 'FORGE_ATTRIBUTE_NAME': (920, 350, 180, 73)}


# 产生一个简单的蜂鸣声
frequency = 1000  # 频率（赫兹）
duration = 1000   # 持续时间（毫秒）


class FORGE(object):
    def __init__(self, base_path1):
        self.base_path1 = base_path1

    def ocr(self):
        ocr_results1, result_details = ocr_paddleocr((save_path1))
        # ocr_results2, result_details = ocr_paddleocr((save_path2))
        corrected_text2 = correct_ocr_errors(ocr_results1, corrections)
        con_ocr_results = f'{corrected_text2}'
        if con_ocr_results in ['部队伤害输出+32%', '战舰能量+42%', '战舰能量+17%']:
            winsound.Beep(frequency, duration)
        print(con_ocr_results)
        append_to_excel(excel_path, con_ocr_results)

    def recycling(self):
        click.center_click(image_names['temple_button'])
        time.sleep(0.1)
        screenshot_region(region_names['FORGE_STATUE_NAME'], save_path1)
        screenshot_region(region_names['FORGE_ATTRIBUTE_NAME'], save_path2)
        click.center_click(image_names['recycling_statues'])
        click.center_click(image_names['recycling'])

    def GREEN_STATUE_S(self, ocr=False):
        click.center_click(image_names['temple_button'])
        click.locate_click(image_names['forge_statue'], offset_x=-248, offset_y=385)
        FORGE.recycling()
        if ocr:
            FORGE.ocr()

    def GREEN_STATUE_M(self, ocr=False):
        click.center_click(image_names['temple_button'])
        click.locate_click(image_names['forge_statue'], offset_x=80, offset_y=385)
        FORGE.recycling()
        if ocr:
            FORGE.ocr()

    def GREEN_STATUE_L(self, ocr=False):
        click.center_click(image_names['temple_button'])
        click.locate_click(image_names['forge_statue'], offset_x=460, offset_y=385)
        FORGE.recycling()
        if ocr:
            FORGE.ocr()

    def BULE_STATUE_S(self, ocr=False):
        click.center_click(image_names['temple_button'])
        click.locate_click(image_names['forge_statue'], offset_x=0, offset_y=72)
        click.locate_click(image_names['forge_statue'], offset_x=-248, offset_y=385)
        FORGE.recycling()
        if ocr:
            FORGE.ocr()

    def BULE_STATUE_M(self, ocr=False):
        click.center_click(image_names['temple_button'])
        click.locate_click(image_names['forge_statue'], offset_x=0, offset_y=72)
        click.locate_click(image_names['forge_statue'], offset_x=80, offset_y=385)
        FORGE.recycling()
        if ocr:
            FORGE.ocr()

    def BULE_STATUE_L(self, ocr=False):
        click.center_click(image_names['temple_button'])
        click.locate_click(image_names['forge_statue'], offset_x=0, offset_y=72)
        click.locate_click(image_names['forge_statue'], offset_x=460, offset_y=385)
        FORGE.recycling()
        if ocr:
            FORGE.ocr()

    def RED_STATUE_S(self, ocr=False):
        click.center_click(image_names['temple_button'])
        click.locate_click(image_names['forge_statue'], offset_x=224, offset_y=72)
        click.locate_click(image_names['forge_statue'], offset_x=-248, offset_y=385)
        FORGE.recycling()
        if ocr:
            FORGE.ocr()

    def RED_STATUE_M(self, ocr=False):
        click.center_click(image_names['temple_button'])
        click.locate_click(image_names['forge_statue'], offset_x=224, offset_y=72)
        click.locate_click(image_names['forge_statue'], offset_x=80, offset_y=385)
        FORGE.recycling()
        if ocr:
            FORGE.ocr()

    def RED_STATUE_L(self, ocr=False):
        click.center_click(image_names['temple_button'])
        click.locate_click(image_names['forge_statue'], offset_x=224, offset_y=72)
        click.locate_click(image_names['forge_statue'], offset_x=460, offset_y=385)
        FORGE.recycling()
        if ocr:
            FORGE.ocr()

    def PURPLE_STATUE_S(self, ocr=False):
        click.center_click(image_names['temple_button'])
        click.locate_click(image_names['forge_statue'], offset_x=458, offset_y=80)
        click.locate_click(image_names['forge_statue'], offset_x=-248, offset_y=385)
        FORGE.recycling()
        if ocr:
            FORGE.ocr()

    def PURPLE_STATUE_M(self, ocr=False):
        click.center_click(image_names['temple_button'])
        click.locate_click(image_names['forge_statue'], offset_x=458, offset_y=80)
        click.locate_click(image_names['forge_statue'], offset_x=80, offset_y=385)
        FORGE.recycling()
        if ocr:
            FORGE.ocr()

    def PURPLE_STATUE_L(self, ocr=False):
        click.center_click(image_names['temple_button'])
        click.locate_click(image_names['forge_statue'], offset_x=458, offset_y=80)
        click.locate_click(image_names['forge_statue'], offset_x=460, offset_y=385)
        FORGE.recycling()
        if ocr:
            FORGE.ocr()


class resarch(object):
    def __init__(self, base_path):
        self.base_path = base_path

    def ocr(self):
        ocr_results1, result_details = ocr_paddleocr(save_path1)
        ocr_results2, result_details = ocr_paddleocr(save_path2)
        if '品质提升' not in ocr_results2:
            ocr_results2 = '-'
        else:
            ocr_results2 = '+'
        combined_results = f"{ocr_results1}{ocr_results2}"
        print(combined_results)
        append_to_excel(excel_path, combined_results)
        click.locate_click(image_names['quality'], offset_x=160, offset_y=195)

    def GREEN_DIAOKEJI(self):
        click.locate_click(image_names['body'], offset_x=188, offset_y=655)
        click.locate_click(image_names['accelerate'], offset_x=156, offset_y=143)
        click.center_click(image_names['accelerate_reserach'])
        click.locate_click(image_names['accelerate_reserach1'], offset_x=350, offset_y=400)
        click.locate_click(image_names['get_award'], offset_x=142, offset_y=522)
        click.double_click(image_names['double_click'])
        screenshot_region(region_names['GREEN_DIAOKEJI_NAME'], save_path1)
        screenshot_region(region_names['GREEN_QUALITY_NAME'], save_path2)
        resarch.ocr()

    def BULE_DIAOKEJI(self):
        click.locate_click(image_names['body'], offset_x=480, offset_y=655)
        click.locate_click(image_names['accelerate'], offset_x=156, offset_y=143)
        click.center_click(image_names['accelerate_reserach'])
        click.locate_click(image_names['accelerate_reserach1'], offset_x=350, offset_y=400)
        click.locate_click(image_names['get_award'], offset_x=142, offset_y=522)
        click.double_click(image_names['double_click'])
        screenshot_region(region_names['BULE_DIAOKEJI_NAME'], save_path1)
        screenshot_region(region_names['BULE_QUALITY_NAME'], save_path2)
        resarch.ocr()

    def RED_DIAOKEJI(self):
        click.locate_click(image_names['body'], offset_x=780, offset_y=655)
        click.locate_click(image_names['accelerate'], offset_x=156, offset_y=143)
        click.center_click(image_names['accelerate_reserach'])
        click.locate_click(image_names['accelerate_reserach1'], offset_x=350, offset_y=400)
        click.locate_click(image_names['get_award'], offset_x=142, offset_y=522)
        click.double_click(image_names['double_click'])
        screenshot_region(region_names['RED_DIAOKEJI_NAME'], save_path1)
        screenshot_region(region_names['RED_QUALITY_NAME'], save_path2)
        resarch.ocr()

    def PURPLE_DIAOKEJI(self):
        click.locate_click(image_names['body'], offset_x=1060, offset_y=655)
        click.locate_click(image_names['accelerate'], offset_x=156, offset_y=143)
        click.center_click(image_names['accelerate_reserach'])
        click.locate_click(image_names['accelerate_reserach1'], offset_x=350, offset_y=400)
        click.locate_click(image_names['get_award'], offset_x=142, offset_y=522)
        click.double_click(image_names['double_click'])
        screenshot_region(region_names['PURPLE_DIAOKEJI_NAME'], save_path1)
        screenshot_region(region_names['PURPLE_QUALITY_NAME'], save_path2)
        resarch.ocr()


class click(object):
    def __init__(self, base_path):
        self.base_path = base_path

    def center_click(self, image_name):
        image_path = os.path.join(self.base_path, image_name)
        if not os.path.exists(image_path):
            print(f"File '{image_path}' does not exist.")
            return
        try:
            img = Image.open(image_path)
            img.verify()  # 验证图像文件的完整性
        except (IOError, SyntaxError) as e:
            print(f"Invalid image file '{image_path}': {e}")
            return
        location = wait_for_image(image_path, timeout=30, confidence=0.8)
        if location:
            center_x, center_y = pyautogui.center(location)
            pyautogui.moveTo(center_x, center_y)
            pyautogui.click()
        else:
            print('点击失败')

    def locate_click(self, image_name, offset_x=0, offset_y=0):
        image_path = os.path.join(self.base_path, image_name)
        if not os.path.exists(image_path):
            print(f"File '{image_path}' does not exist.")
            return
        try:
            img = Image.open(image_path)
            img.verify()  # 验证图像文件的完整性
        except (IOError, SyntaxError) as e:
            print(f"Invalid image file '{image_path}': {e}")
            return
        location = wait_for_image(image_path, timeout=30, confidence=0.8)
        if location:
            left, top = location.left, location.top
            click_x = left + offset_x
            click_y = top + offset_y
            pyautogui.moveTo(click_x, click_y)
            pyautogui.click()
        else:
            print('点击失败')

    def double_click(self, image_name):
        image_path = os.path.join(self.base_path, image_name)
        location = wait_for_image(image_path, timeout=30, confidence=0.8)
        if location:
            center_x, center_y = pyautogui.center(location)
            pyautogui.moveTo(center_x, center_y)
            pyautogui.click()
            pyautogui.click()
        else:
            print('点击失败')


def ocr_paddleocr(image_path):
    # 创建一个 PaddleOCR 对象，指定语言为中文（简体）和英文

    args = {
        'use_gpu': False,
        'gpu_id': 0,  # 这里设置为1，表示使用第二块GPU
        'det_algorithm': 'DB',
        'rec_algorithm': 'SVTR_LCNet',
        'gpu_mem': 500,
        'det': True,
        'rec': True,
        'rec_batch_num': 32,  # 增加批处理大小
        'cls_batch_num': 32,  # 增加批处理大小
        'det_model_dir': r'C:\Users\zzzzz\.paddleocr\whl\det\ch\ch_PP-OCRv4_det_infer',
        'rec_model_dir': r'C:\Users\zzzzz\.paddleocr\whl\rec\ch\ch_PP-OCRv4_rec_infer',
        'cls_model_dir': r'C:\Users\zzzzz\.paddleocr\whl\cls\ch_ppocr_mobile_v2.0_cls_infer'
    }

    ocr = PaddleOCR(**args)

    # 读取图像文件并进行 OCR 识别
    result = ocr.ocr(image_path, cls=True)

    # 提取识别的文字并拼接成字符串
    text = "".join([line[1][0] for line in result[0]])

    return text, result

def correct_ocr_errors(text, corrections):
    """
    使用提供的字典 corrections 来校正 OCR 输出的错误
    :param text: OCR 输出的文本
    :param corrections: 包含错误和正确值的字典
    :return: 校正后的文本
    """
    for incorrect, correct in corrections.items():
        text = text.replace(incorrect, correct)
    return text

# 错误校正字典
corrections = {
    "%九1+": "+14%",
    "%功1+": "+14%",
    "%h1+": "+14%",
}


def visualize_ocr(image_path, result):
    image = Image.open(image_path).convert('RGB')
    boxes = [line[0] for line in result[0]]
    txts = [line[1][0] for line in result[0]]
    scores = [line[1][1] for line in result[0]]

    im_show = draw_ocr(image, boxes, txts, scores, font_path='path_to_font.ttf')
    im_show = Image.fromarray(im_show)
    im_show.save('result.jpg')


def screenshot_region(region, save_path):
    """
    截取屏幕上的特定区域并保存为图像文件。

    :param region: 截图区域的坐标和大小 (left, top, width, height)
    :param save_path: 保存图像的路径
    """
    # 截取屏幕指定区域
    screenshot = pyautogui.screenshot(region=region)

    # 保存截图
    screenshot.save(save_path)
    # print(f"截图已保存至 {save_path}")


def wait_for_image(image_path, timeout=30, confidence=0.8):
    """
    等待图像在屏幕上显示。

    :param image_path: 图像文件路径
    :param timeout: 最大等待时间（秒）
    :param confidence: 置信度
    :return: 图像位置（如果找到）或 None
    """
    start_time = time.time()
    while time.time() - start_time < timeout:
        try:
            location = pyautogui.locateOnScreen(image_path, confidence=confidence)
            if location:
                return location
        except pyautogui.ImageNotFoundException:
            pass
        time.sleep(0.001)
    return None


def append_to_excel(excel_path, data):
    # 尝试打开现有的工作簿，如果不存在则创建一个新的
    try:
        wb = load_workbook(excel_path)
        ws = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active

    # 找到最后一行并在下一行写入数据
    ws.append([data])

    # 保存Excel文件
    wb.save(excel_path)


#寻找雕刻机并点击研究
def run_DIAOKEJI():
    click.center_click(image_names['LEIDIAN'])
    click.center_click(image_names['DIAOKEJI'])
    click.center_click(image_names['reserch'])
    click.center_click(image_names['study'])


# 神庙雕像序列研究
def run_TEMPLE():
    click.center_click(image_names['LEIDIAN'])
    click.center_click(image_names['temple'])

click = click(base_path1)
FORGE = FORGE(base_path1)
resarch = resarch(base_path1)